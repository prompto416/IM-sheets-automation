import sys 
#directory for modules
sys.path.append("\modules")

import gspread
from oauth2client.service_account import ServiceAccountCredentials  
import time
from scrape import *
from getDownloadURL_Module import *
from initiateDownload import *
import csv
import datetime

scope =["https://spreadsheets.google.com/feeds",'https://www.googleapis.com/auth/spreadsheets',"https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name('testSheetAPI_Key.json', scope)
client = gspread.authorize(creds)

#python_test = client.open('testSheet').sheet1
python_test = client.open('testSheet').worksheet('IM')

#if outdated: then update
myHTML = scrapeHTML_string()
downloadURL = getDownloadURL(myHTML)
fileForUpdate = getDownloadURL_fileName(myHTML)
#check for update before updating
#mind the file name
checkFile = open('Last Version Note.txt','r')
lastVersion = checkFile.readline().strip()
if lastVersion == fileForUpdate:
    print('The file is up to date.')
    #quit()
else:
    print('Initiating Update.')
    newFile = open('Last Version Note.txt','w')
    newFile.write(fileForUpdate)
    
    

initDownload(downloadURL)




from openpyxl import load_workbook
#fileForUpdate = string ชื่อไฟล์
# wb = ชื่อไฟล์ที่จะใช้อ้างอิง update 

# fileForUpdate = 'Margin_Announcement_20220628.xlsx'
wb = load_workbook(filename=fileForUpdate, 
read_only=True)

#declaration of sheet page
ws = wb['margin_Noti']

# Read the cell values into a list of lists
data_rows = []
data_rows_val = []
for row in ws['B37':'B164']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

for row in ws['E37':'E164']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows_val.append(data_cols)
# Transform into dataframe
import pandas as pd
df = pd.DataFrame(data_rows)

# rows of nonStock symbols
# b13 d
# b18 d
# b19 d
# b21 d
# b25 d
# b22 d
# b34 d


print(ws['B13'].value,True,int(ws['D13'].value)*1.75)
print(ws['B18'].value,True)
print(ws['B19'].value,True)
print(ws['B21'].value,True)
print(ws['B25'].value,True)
print(ws['B22'].value,True)
print(ws['B34'].value,True)

for i in data_rows:
    for j in i:
        if j == 'SCBB':
            print(data_rows.index(i))
            toBeRemoved = data_rows.index(i)
print(toBeRemoved)
print(len(data_rows))
print(data_rows[90])
print("before: ",len(data_rows),len(data_rows_val))
data_rows.pop(toBeRemoved)
data_rows_val.pop(toBeRemoved)

print(len(data_rows),len(data_rows_val))
#print(data_rows)
#print(data_rows_val)
nonStockBase = client.open('testSheet').worksheet('Non-Stock Market Base')
def nonStock_new2old():
    nonStockListIndex = [13,18,19,21,25,22,34]
    for i in range(7):
        IM_new = nonStockBase.acell('C'+str(2+i)).value
        nonStockBase.update('D'+str(2+i),IM_new)
        nonStockBase.update('C'+str(2+i),ws['D'+str(nonStockListIndex[i])].value*1.75)
        
        


def updateSheet():
    for i in range(len(data_rows)):
        python_test.update_cell(2+i,2,data_rows[i][0])
        python_test.update_cell(2+i,3,data_rows_val[i][0])
        if (i!=0) and (i%29)==0:
            print('sleeping for a minute')
            time.sleep(60)
        print('writing iteration:',i)
current = datetime.datetime.now()
now = current.strftime('%a %y/%m/%d')
python_test.update_cell(1,15,f'อัพเดทล่าสุดวันที่: {now} ชื่อไฟล์: {fileForUpdate}')
def backupSheet():
    docid = '1WiWQfKX1ACVSIHWszrnlVtpQ_eZUqOqIkA1xAk7ccRY'
    backup = client.open_by_key(docid)
    for i, worksheet in enumerate(backup.worksheets()):
        filename = docid + '-worksheet' + str(i) + '.csv'
        with open(filename, 'wb') as f:
            writer = csv.writer(f)
            writer.writerows(worksheet.get_all_values())
        
#backupSheet()
def downloadPDF():
    spreadsheet_name = 'testSheet'
    spreadsheet = client.open(spreadsheet_name)
    url = 'https://docs.google.com/spreadsheets/export?format=pdf&id=' + spreadsheet.id
    headers = {'Authorization': 'Bearer ' + creds.create_delegated("").get_access_token().access_token}
    res = requests.get(url, headers=headers)
    with open(spreadsheet_name + ".pdf", 'wb') as f:
        f.write(res.content)


#updateSheet()
print(nonStockBase.acell('E2').value)
print(type(nonStockBase.acell('E2').value))
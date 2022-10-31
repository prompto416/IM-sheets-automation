from openpyxl import load_workbook

wb = load_workbook(filename='Margin_Announcement_20220628.xlsx', 
                   read_only=True)

ws = wb['margin_Noti']

# Read the cell values into a list of lists
data_rows = []
data_rows_val = []
for row in ws['B37':'B164']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

for row in ws['E37':'E164']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows_val.append(data_cols)
# Transform into dataframe
import pandas as pd
df = pd.DataFrame(data_rows)


for i in range(len(data_rows)):
    print(data_rows[i][0],data_rows_val[i][0])
    




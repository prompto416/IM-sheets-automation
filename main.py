import gspread
from oauth2client.service_account import ServiceAccountCredentials  
import time
import pandas as pd
from openpyxl import load_workbook
import datetime
from scrape import *
from getDownloadURL_Module import *
from initiateDownload import *
import datetime
# from checkDate import *
now = datetime.datetime.now()

current_time = now.strftime("%H:%M:%S")
current_date = now.strftime("%a %d/%m/%y")
print(current_date,current_time)

log_file = open('log.txt','a')

scope =["https://spreadsheets.google.com/feeds",'https://www.googleapis.com/auth/spreadsheets',"https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"]
#your API KEY
creds = ServiceAccountCredentials.from_json_keyfile_name('', scope)
client = gspread.authorize(creds)

#your sheet name
python_test = client.open().sheet1
# python_test = client.open("deleteme").sheet1



#if outdated: then update
myHTML = scrapeHTML_string()
htmlSET = open("htmlSET.txt","w")
htmlSET.write(myHTML)
downloadURL = getDownloadURL(myHTML)
fileForUpdate = getDownloadURL_fileName(myHTML)
initDownload(downloadURL)


#check for update before updating
#mind the file name
checkFile = open('cache.txt','r')
lastVersion = checkFile.readline().strip()
if lastVersion == fileForUpdate:
    print('The file is up to date.')
    log_file.write(f'{current_date} {current_time} = No Update')
    log_file.write('\n')
    quit()
else:
    print('Initiating Update.')
    try: 
        log_file.write(f'{current_date} {current_time} = Updating with file: {fileForUpdate} URL: {downloadURL}')
        log_file.write('\n')
    except:
        log_file.write(f'{current_date} {current_time} = Error!')
        log_file.write('\n')
   
    
#fileForUpdate = string ชื่อไฟล์
# wb = ชื่อไฟล์ที่จะใช้อ้างอิง update 

# fileForUpdate = 'Margin_Announcement_20220628.xlsx'
wb = load_workbook(filename=fileForUpdate, 
read_only=True)

ws = wb['margin_Noti']

# Read the cell values into a list of lists
data_rows = []
data_rows_val = []
for row in ws['B37':'B164']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

for row in ws['D37':'D164']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows_val.append(data_cols)
# Transform into dataframe
# df = pd.DataFrame(data_rows)

for i in data_rows:
    for j in i:
        if j == 'SCBB':
            print(data_rows.index(i))
            toBeRemoved = data_rows.index(i)
print(toBeRemoved)
print(len(data_rows))
print(data_rows[90])
print("before: ",len(data_rows),len(data_rows_val))
data_rows.pop(toBeRemoved)
data_rows_val.pop(toBeRemoved)

#copy and paste IM ใหม่ ใส่ IM เดิม

        
nonStockBase = client.open('Line_bot_IM (New Update)').worksheet('Non-Stock Market Base')
def nonStock_new2old():
    
    #Rows of nonstock in excel
    nonStockListIndex = [13,18,19,21,25,22,34]
    for i in range(7):
        IM_new = nonStockBase.acell('C'+str(2+i)).value
        nonStockBase.update('D'+str(2+i),IM_new)
        nonStockBase.update('C'+str(2+i),ws['D'+str(nonStockListIndex[i])].value*1.75)
#ดึงข้อมูลจากเบส nonstock ให้ไฟล์ editImage.py 


def updateSheet():
    print(data_rows)
    print(data_rows_val)
    python_test.batch_update([{
        'range': 'B2:B999',
        'values': data_rows,
        }, ])
    python_test.batch_update([{
        'range': 'C2:C999',
        'values': data_rows_val,
        }, ])
    
    #โค้ดเก่า Linear Complexity 
    # for i in range(len(data_rows)):
    #     python_test.update_cell(2+i,2,data_rows[i][0])
    #     python_test.update_cell(2+i,3,data_rows_val[i][0])
    #     if (i!=0) and (i%29)==0:
    #         print('sleeping for a minute')
    #         time.sleep(60)
    #     print('writing iteration:',i)
        
def backupSheet():
    docid = '1WiWQfKX1ACVSIHWszrnlVtpQ_eZUqOqIkA1xAk7ccRY'
    for i, worksheet in enumerate(python_test.worksheets()):
        filename = docid + '-worksheet' + str(i) + '.csv'
        with open(filename, 'wb') as f:
            writer = csv.writer(f)
            writer.writerows(worksheet.get_all_values())

def downloadPDF():
    spreadsheet_name = 'Line_bot_IM (New Update)'
    spreadsheet = client.open(spreadsheet_name)
    url = 'https://docs.google.com/spreadsheets/export?format=pdf&id=' + spreadsheet.id
    headers = {'Authorization': 'Bearer ' + creds.create_delegated("").get_access_token().access_token}
    res = requests.get(url, headers=headers)
    with open(spreadsheet_name + ".pdf", 'wb') as f:
        f.write(res.content)
       
#nonStock_new2old()

# IM_newRow = []
# IM_oldRow = []
# IM_changeRow = []      
# for i in range(7):
#     IM_newRow.append(nonStockBase.acell('C'+str(2+i)).value)
#     IM_oldRow.append(nonStockBase.acell('D'+str(2+i)).value)
#     IM_changeRow.append(nonStockBase.acell('E'+str(2+i)).value)
# print(IM_newRow)
# print(IM_oldRow)
# print(IM_changeRow)

updateSheet()

#Saving the lastest version after updating to prevent saving without updating
newFile = open('cache.txt','w')
newFile.write(fileForUpdate)



